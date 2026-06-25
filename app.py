from flask import Flask, render_template, request, jsonify, send_file, session, redirect, url_for
import os
import asyncio
import json
from datetime import datetime, date
from dotenv import load_dotenv
from pipeline import executar_pipeline, executar_pipeline_selecionados
from scraper import coletar_links_artigos
 
# 🔌 Carregar .env APENAS localmente
if os.environ.get("RENDER") is None:
    load_dotenv()
 
# 🔐 Função para pegar senha (SECRET FILE + fallback ENV)
def get_password():
    # 1. Tenta SECRET FILE (Render)
    secret_path = "/etc/secrets/APP_PASSWORD"
 
    if os.path.exists(secret_path):
        try:
            with open(secret_path) as f:
                senha = f.read().strip()
                print("✅ Senha carregada via SECRET FILE")
                return senha
        except Exception as e:
            print("❌ Erro ao ler secret file:", e)
 
    # 2. Fallback: variável de ambiente
    senha_env = os.getenv("APP_PASSWORD")
    if senha_env:
        print("✅ Senha carregada via ENV")
        return senha_env
 
    # 3. Falhou tudo
    print("🚨 Nenhuma senha encontrada!")
    return None
 
 
# 🔐 Variáveis do sistema
APP_PASSWORD = get_password()
FLASK_SECRET_KEY = os.getenv("FLASK_SECRET_KEY", "chave_super_secreta_padrao")
 
app = Flask(__name__)
app.secret_key = FLASK_SECRET_KEY
 
# 🌍 Mapa de países
MAPA_PAISES = {
    'pt_br': 'Brasil', 'en_us': 'Estados Unidos', 'en_ca': 'Canadá',
    'en_au': 'Austrália e Nova Zelândia', 'en_za': 'África do Sul', 'en_gb': 'Reino Unido',
    'es_es': 'Espanha', 'es_mx': 'México', 'fr_fr': 'França', 'nl_nl': 'Holanda',
    'en_eu': 'Europa (outros países)', 'en_ap': 'Ásia (outros países)',
    'en_la': 'América Latina (outros países)', 'es_la': 'América Latina (outros países)',
    'de_de': 'Alemanha', 'it_it': 'Itália', 'ja_jp': 'Japão', 'zh_cn': 'China', 'pt_pt': 'Portugal'
}
 
# 📂 Pasta de resultados
PASTA_RESULTADOS = os.path.join(os.getcwd(), "resultados")
os.makedirs(PASTA_RESULTADOS, exist_ok=True)
 
print("📁 Pasta de resultados:", PASTA_RESULTADOS)
 
# 🔐 LOGIN
@app.route("/", methods=["GET", "POST"])
def login():
    erro = None
 
    if request.method == "POST":
        senha = request.form.get("password")
 
        print("🔎 DEBUG - senha digitada:", "OK" if senha else "VAZIA")
        print("🔎 DEBUG - senha sistema:", "OK" if APP_PASSWORD else "NÃO CARREGADA")
 
        if not APP_PASSWORD:
            print("🚨 ERRO CRÍTICO: senha não carregada")
            erro = "Erro interno no servidor"
            return render_template("login.html", erro=erro)
 
        if senha and senha.strip() == APP_PASSWORD.strip():
            print("✅ LOGIN OK")
            session["logado"] = True
            return redirect(url_for("home"))
        else:
            print("❌ LOGIN FALHOU")
            erro = "Senha incorreta"
 
    return render_template("login.html", erro=erro)
 
 
# 🏠 Página protegida
@app.route("/home")
def home():
    if not session.get("logado"):
        return redirect(url_for("login"))
    return render_template("index.html")
 
 
# 📘 Manual
@app.route("/manual")
def manual():
    if not session.get("logado"):
        return redirect(url_for("login"))
    return render_template("manual.html")
 
 
# 🌍 API países
@app.route("/paises", methods=["GET"])
def get_paises():
    if not session.get("logado"):
        return jsonify({"erro": "Não autorizado"}), 401
    return jsonify(MAPA_PAISES)
 
 
# 🚀 Rodar scraper
@app.route("/rodar", methods=["POST"])
def rodar():
    if not session.get("logado"):
        return jsonify({"erro": "Não autorizado"}), 401
 
    try:
        data = request.get_json()
        if not data:
            return jsonify({"sucesso": False, "erro": "Nenhum dado recebido"}), 400
 
        pais = data.get("pais")
        quantidade = int(data.get("quantidade", 1))
 
        if pais not in MAPA_PAISES:
            return jsonify({"sucesso": False, "erro": "País inválido"}), 400
 
        print(f"🌍 País: {pais} | 📄 Quantidade: {quantidade}")
 
        status, arquivos = asyncio.run(
            executar_pipeline(
                pais_input=pais,
                alias_input=None,
                qtd_artigos=quantidade
            )
        )
 
        print("📂 Arquivos gerados:", arquivos)
 
        _enfileirar_notificacao(
            "sucesso",
            "Tradução concluída!",
            f"{len(arquivos)} arquivo(s) gerado(s) para {MAPA_PAISES.get(pais, pais)}."
        )
        return jsonify({
            "sucesso": True,
            "status": status,
            "arquivos": arquivos
        })
 
    except Exception as e:
        print("❌ Erro no /rodar:", str(e))
        _enfileirar_notificacao("erro", "Erro na tradução", str(e)[:120])
        return jsonify({"sucesso": False, "erro": str(e)}), 500
 
 
# 📥 Download
@app.route("/download/<path:nome_arquivo>", methods=["GET"])
def download(nome_arquivo):
    if not session.get("logado"):
        return jsonify({"erro": "Não autorizado"}), 401
 
    caminho = os.path.join(PASTA_RESULTADOS, nome_arquivo)
    print("📥 Download:", caminho)
 
    if os.path.exists(caminho):
        return send_file(caminho, as_attachment=True)
    else:
        print("❌ Arquivo não encontrado")
        return jsonify({"erro": "Arquivo não encontrado"}), 404
 
 
# 🔍 Página de pesquisa de artigos
@app.route("/pesquisar")
def pesquisar():
    if not session.get("logado"):
        return redirect(url_for("login"))
    return render_template("pesquisar.html")
 
 
# 🔍 API: buscar lista de artigos disponíveis
@app.route("/buscar-artigos", methods=["POST"])
def buscar_artigos():
    if not session.get("logado"):
        return jsonify({"erro": "Não autorizado"}), 401
 
    try:
        data = request.get_json()
        pais = data.get("pais")
        busca = data.get("busca", "").strip().lower()
 
        if pais not in MAPA_PAISES:
            return jsonify({"sucesso": False, "erro": "País inválido"}), 400
 
        url_blog = f"https://www.tennantco.com/{pais}/blog.html"
        links = coletar_links_artigos(url_blog, pais)
 
        if busca:
            links = [l for l in links if busca in l.get("title", "").lower() or busca in l.get("href", "").lower()]
 
        return jsonify({
            "sucesso": True,
            "artigos": links,
            "total": len(links),
            "pais_nome": MAPA_PAISES.get(pais, pais)
        })
 
    except Exception as e:
        print("❌ Erro no /buscar-artigos:", str(e))
        return jsonify({"sucesso": False, "erro": str(e)}), 500
 
 
# 🚀 Traduzir artigos selecionados
@app.route("/traduzir-selecionados", methods=["POST"])
def traduzir_selecionados():
    if not session.get("logado"):
        return jsonify({"erro": "Não autorizado"}), 401
 
    try:
        data = request.get_json()
        pais = data.get("pais")
        urls_selecionadas = data.get("urls", [])
 
        if pais not in MAPA_PAISES:
            return jsonify({"sucesso": False, "erro": "País inválido"}), 400
 
        if not urls_selecionadas:
            return jsonify({"sucesso": False, "erro": "Nenhum artigo selecionado"}), 400
 
        print(f"🎯 Traduzindo {len(urls_selecionadas)} artigos selecionados do país {pais}")
 
        status, arquivos = asyncio.run(
            executar_pipeline_selecionados(
                pais_input=pais,
                urls_selecionadas=urls_selecionadas
            )
        )
 
        _enfileirar_notificacao(
            "sucesso",
            "Artigos traduzidos!",
            f"{len(arquivos)} arquivo(s) gerado(s)."
        )
        return jsonify({
            "sucesso": True,
            "status": status,
            "arquivos": arquivos
        })
 
    except Exception as e:
        print("❌ Erro no /traduzir-selecionados:", str(e))
        _enfileirar_notificacao("erro", "Erro ao traduzir selecionados", str(e)[:120])
        return jsonify({"sucesso": False, "erro": str(e)}), 500
 
 
# 🤖 TENNIX — Assistente Inteligente
@app.route("/tennix")
def tennix():
    if not session.get("logado"):
        return redirect(url_for("login"))
    return render_template("tennix.html")
 
 
# 🤖 TENNIX — Chat endpoint com IA (Gemini)
@app.route("/tennix-chat", methods=["POST"])
def tennix_chat():
    if not session.get("logado"):
        return jsonify({"erro": "Não autorizado"}), 401
 
    try:
        import requests as req_lib
 
        data = request.get_json()
        mensagem     = data.get("mensagem", "").strip()
        historico    = data.get("historico", [])
        pais         = data.get("pais", "")
        pais_nome    = data.get("pais_nome", "")
        artigos      = data.get("artigos", [])
        imagem_b64   = data.get("imagem_base64", None)
        imagem_type  = data.get("imagem_media_type", "image/jpeg")
 
        if not mensagem:
            return jsonify({"erro": "Mensagem vazia"}), 400
 
        # ── System prompt da TENNIX ──
        artigos_str = ""
        if artigos:
            linhas = [f"{i+1}. [{a.get('title','s/título')}] ({a.get('href','')})"
                      for i, a in enumerate(artigos[:200])]
            artigos_str = "\n".join(linhas)
        else:
            artigos_str = "(nenhum artigo carregado ainda)"
 
        system_prompt = f"""Você é TENNIX, assistente inteligente da TENNIX by Tennant Company.
Você ajuda a encontrar artigos dos blogs globais Tennant com base em temas, independente do idioma dos títulos.
 
PAÍS ATUAL: {pais_nome or 'não selecionado'} ({pais or '-'})
 
ARTIGOS DISPONÍVEIS ({len(artigos)} total):
{artigos_str}
 
INSTRUÇÕES:
- Responda SEMPRE em português brasileiro, de forma amigável e profissional.
- Quando o usuário pedir artigos sobre um tema, analise os títulos acima (em qualquer idioma) e identifique quais são relevantes.
- Se o usuário pedir "todos os artigos", retorne todos os URLs disponíveis.
- Se não houver artigos carregados, oriente o usuário a selecionar um país e carregar os artigos.
 
REGRA CRÍTICA DE FORMATO:
Sua resposta deve ser EXCLUSIVAMENTE um objeto JSON válido. Nada antes, nada depois.
NÃO escreva texto livre. NÃO use markdown. NÃO use ```. APENAS o JSON abaixo:
 
{{"texto": "sua resposta amigável em português aqui", "acao": null, "artigos_filtrados": ["url1", "url2"], "tipo_filtro": null}}
 
Campos obrigatórios:
- "texto": string com resposta clara em português (ex: "Encontrei 6 artigos sobre limpeza industrial:")
- "artigos_filtrados": array de URLs relevantes, ou [] se nenhum
- "acao": null, ou "selecionar_todos" se o usuário quiser todos os artigos
- "tipo_filtro": null, ou um dos valores "produto","dica","case","feriado","conteudo","outro" se o usuário filtrar por tipo
"""
 
        # ── Monta histórico para a OpenAI ──
        messages = [{"role": "system", "content": system_prompt}]
        for h in historico[-8:]:
            role = h.get("role", "")
            content_h = h.get("content", "")
            if not content_h or role not in ("user", "assistant"):
                continue
            messages.append({"role": role, "content": content_h})
 
        # Adiciona a mensagem atual (com imagem se houver)
        if imagem_b64:
            messages.append({
                "role": "user",
                "content": [
                    {
                        "type": "image_url",
                        "image_url": {
                            "url": f"data:{imagem_type};base64,{imagem_b64}",
                            "detail": "low"
                        }
                    },
                    {"type": "text", "text": mensagem}
                ]
            })
        else:
            messages.append({"role": "user", "content": mensagem})
 
        # ── Chama a API da OpenAI ──
        api_key = os.getenv("OPENAI_API_KEY", "").strip()
        print(f"🔑 OPENAI_API_KEY: {'OK' if api_key else 'NÃO CONFIGURADA'}")
 
        if not api_key:
            return jsonify({
                "texto": "⚠️ Chave de API da TENNIX não configurada. Configure a variável OPENAI_API_KEY no Render.",
                "artigos_filtrados": [],
                "acao": None
            })
 
        resp = req_lib.post(
            "https://api.openai.com/v1/chat/completions",
            headers={
                "Content-Type": "application/json",
                "Authorization": f"Bearer {api_key}"
            },
            json={
                "model": "gpt-4o-mini",
                "max_tokens": 2000,
                "temperature": 0.2,
                "response_format": {"type": "json_object"},
                "messages": messages
            },
            timeout=30
        )
        if not resp.ok:
            print("❌ OpenAI error:", resp.status_code, resp.text[:300])
        resp.raise_for_status()
        result = resp.json()
 
        # ── Extrai texto da resposta da OpenAI ──
        raw_text = ""
        try:
            raw_text = result["choices"][0]["message"]["content"]
        except Exception as e:
            print("❌ Erro ao extrair texto da OpenAI:", e)
 
        print("📨 OpenAI raw_text:", raw_text[:300])
 
        # ── Parse JSON robusto ──
        # Tenta várias estratégias para extrair o JSON mesmo que o modelo
        # coloque texto livre antes ou depois
        def extrair_json(text):
            import re
            text = text.strip()
 
            # 1. Tenta direto (ideal — response_format funcionou)
            try:
                return json.loads(text)
            except Exception:
                pass
 
            # 2. Remove blocos ```json ... ``` ou ``` ... ```
            text_sem_md = re.sub(r'```json\s*', '', text)
            text_sem_md = re.sub(r'```\s*', '', text_sem_md).strip()
            try:
                return json.loads(text_sem_md)
            except Exception:
                pass
 
            # 3. Encontra o primeiro { ... } completo no texto (ignora texto livre ao redor)
            match = re.search(r'\{[\s\S]*\}', text)
            if match:
                try:
                    return json.loads(match.group(0))
                except Exception:
                    pass
 
            return None
 
        parsed = extrair_json(raw_text)
 
        if not parsed:
            print("⚠️ Não conseguiu parsear JSON. raw_text:", raw_text[:500])
            parsed = {
                "texto": "Desculpe, tive um problema ao processar a resposta. Tente novamente.",
                "artigos_filtrados": [],
                "acao": None
            }
 
        return jsonify({
            "texto":             parsed.get("texto", ""),
            "artigos_filtrados": parsed.get("artigos_filtrados", []),
            "acao":              parsed.get("acao", None),
            "tipo_filtro":       parsed.get("tipo_filtro", None),
        })
 
    except Exception as e:
        print("❌ Erro no /tennix-chat:", str(e))
        return jsonify({"erro": f"Erro interno: {str(e)}"}), 500


# ─────────────────────────────────────────────────────────────
# 🔍 AUDITORIA — OpenAI GPT-4o com web search
# ─────────────────────────────────────────────────────────────
@app.route("/tennix-auditoria", methods=["POST"])
def tennix_auditoria():
    """
    Endpoint de auditoria que usa OpenAI Responses API (gpt-4o)
    com a ferramenta web_search_preview para buscar documentos
    no site da Tennant de forma segura via backend.
    """
    if not session.get("logado"):
        return jsonify({"erro": "Não autorizado"}), 401

    try:
        import requests as req_lib

        data = request.get_json()
        tipo      = data.get("tipo", "")       # "listar_maquinas" | "verificar_docs"
        maquina   = data.get("maquina", None)  # {"nome": ..., "url": ..., "categoria": ...}
        docs      = data.get("docs", [])        # lista de chaves de documentos

        api_key = os.getenv("OPENAI_API_KEY", "").strip()
        if not api_key:
            return jsonify({"erro": "OPENAI_API_KEY não configurada no servidor."}), 500

        # ── Helpers ──────────────────────────────────────────────
        def chamar_openai(prompt_text):
            """Chama a OpenAI Responses API com web_search_preview."""
            resp = req_lib.post(
                "https://api.openai.com/v1/responses",
                headers={
                    "Content-Type": "application/json",
                    "Authorization": f"Bearer {api_key}"
                },
                json={
                    "model": "gpt-4o",
                    "tools": [{"type": "web_search_preview"}],
                    "input": prompt_text,
                    "max_output_tokens": 3000
                },
                timeout=60
            )
            if not resp.ok:
                raise Exception(f"OpenAI API error {resp.status_code}: {resp.text[:300]}")
            result = resp.json()
            # Extrai texto da resposta
            texto = ""
            for item in result.get("output", []):
                if item.get("type") == "message":
                    for c in item.get("content", []):
                        if c.get("type") == "output_text":
                            texto += c.get("text", "")
            return texto

        def extrair_json_seguro(texto):
            import re
            texto = texto.strip()
            try:
                return json.loads(texto)
            except Exception:
                pass
            texto2 = re.sub(r'```json\s*', '', texto)
            texto2 = re.sub(r'```\s*', '', texto2).strip()
            try:
                return json.loads(texto2)
            except Exception:
                pass
            match = re.search(r'\{[\s\S]*\}', texto)
            if match:
                try:
                    return json.loads(match.group(0))
                except Exception:
                    pass
            return None

        # ── Tipo 1: Listar máquinas ───────────────────────────────
        if tipo == "listar_maquinas":
            categorias_alvo = data.get("categorias", None)  # lista opcional de categorias

            if categorias_alvo:
                cats_str = "\n".join([f"- {c}" for c in categorias_alvo])
                prompt = f"""Acesse o site da Tennant Company Brasil e liste TODOS os modelos de máquinas das seguintes categorias:

{cats_str}

URLs base para verificar:
- https://www.tennantco.com/pt_br/m%C3%A1quinas.html (página principal)
- Para Extratoras de carpete: https://www.tennantco.com/pt_br/m%C3%A1quinas/extratoras-de-carpete.html
- Para Polidoras e enceradeiras: https://www.tennantco.com/pt_br/m%C3%A1quinas/polidoras-e-enceradeiras.html
- Para Aspiradores: https://www.tennantco.com/pt_br/m%C3%A1quinas/aspiradores.html
- Para Equipamento de limpeza especializada: https://www.tennantco.com/pt_br/m%C3%A1quinas/equipamentos-especializados.html
- Para Lavadoras de alta pressão: https://www.tennantco.com/pt_br/m%C3%A1quinas/lavadoras-de-alta-pressao.html

Acesse cada URL acima que corresponda às categorias solicitadas. Para cada modelo individual encontrado forneça:
- nome: modelo da máquina (ex: R3, A5, B5, etc.)
- url: URL completa da página do produto
- categoria: exatamente o nome da categoria conforme listado acima

Responda APENAS com JSON válido, sem texto antes ou depois:
{{"maquinas":[{{"nome":"R3","url":"https://www.tennantco.com/pt_br/...","categoria":"Extratoras de carpete"}},{{"nome":"A5","url":"...","categoria":"Aspiradores"}}]}}"""
            else:
                prompt = """Acesse o site da Tennant Company Brasil e liste TODOS os modelos de máquinas disponíveis no catálogo.

URLs para verificar:
- https://www.tennantco.com/pt_br/m%C3%A1quinas.html (página principal de máquinas)
- Subcategorias: lavadoras de piso, varredeiras, extratoras, polidoras, aspiradores, robóticas
- Para Extratoras de carpete: https://www.tennantco.com/pt_br/m%C3%A1quinas/extratoras-de-carpete.html
- Para Polidoras e enceradeiras: https://www.tennantco.com/pt_br/m%C3%A1quinas/polidoras-e-enceradeiras.html
- Para Aspiradores: https://www.tennantco.com/pt_br/m%C3%A1quinas/aspiradores.html
- Para Equipamento de limpeza especializada: https://www.tennantco.com/pt_br/m%C3%A1quinas/equipamentos-especializados.html
- Para Lavadoras de alta pressão: https://www.tennantco.com/pt_br/m%C3%A1quinas/lavadoras-de-alta-pressao.html

Para cada modelo encontrado, forneça:
- nome: modelo da máquina (ex: T360, T300, B70, A260, S30, etc.)
- url: URL completa da página do produto no site pt_br
- categoria: categoria do produto

Responda APENAS com JSON válido, sem texto antes ou depois:
{"maquinas":[{"nome":"T360","url":"https://www.tennantco.com/pt_br/...","categoria":"Lavadora a pé"},{"nome":"T300","url":"...","categoria":"..."}]}"""

            texto_resp = chamar_openai(prompt)
            parsed = extrair_json_seguro(texto_resp)

            if not parsed or "maquinas" not in parsed:
                # Fallback: tentar extrair qualquer lista
                return jsonify({
                    "sucesso": False,
                    "erro": "Não foi possível obter a lista de máquinas do site.",
                    "raw": texto_resp[:500]
                })

            return jsonify({
                "sucesso": True,
                "maquinas": parsed["maquinas"]
            })

        # ── Tipo 2: Verificar documentos de uma máquina ──────────
        elif tipo == "verificar_docs":
            if not maquina or not docs:
                return jsonify({"erro": "Dados insuficientes: maquina e docs são obrigatórios"}), 400

            DOC_LABELS = {
                "folheto":         "Folheto / Brochure",
                "guia":            "Guia de peças de reposição",
                "manual_pecas":    "Manual de peças",
                "manual_operador": "Manual do operador",
                "tabela_parede":   "Tabela de parede / Wall chart"
            }

            docs_lista = "\n".join([f'- {DOC_LABELS.get(d, d)} (chave: "{d}")' for d in docs])

            prompt = f"""Você é um auditor do site da Tennant Company Brasil.

Verifique se o modelo de máquina "{maquina['nome']}" possui os seguintes documentos disponíveis para download no site tennantco.com:

Documentos a verificar:
{docs_lista}

Como pesquisar:
1. Acesse a página do produto: {maquina.get('url', f'https://www.tennantco.com/pt_br pesquisar {maquina["nome"]}')}
2. Procure pela seção de documentos/downloads da página
3. Também busque por: site:tennantco.com "{maquina['nome']}" filetype:pdf

Para cada documento:
- "sim" = documento encontrado e disponível para download
- "nao" = documento não encontrado

Responda APENAS com JSON válido:
{{"documentos":{{{", ".join([f'"{d}": "sim" ou "nao"' for d in docs])}}}}}

Exemplo de resposta:
{{"documentos":{{{", ".join([f'"{d}": "nao"' for d in docs])}}}}}"""

            texto_resp = chamar_openai(prompt)
            parsed = extrair_json_seguro(texto_resp)

            # Normalizar resultado
            doc_result = {}
            if parsed and "documentos" in parsed:
                for d in docs:
                    val = str(parsed["documentos"].get(d, "nao")).lower()
                    doc_result[d] = "sim" if "sim" in val else "nao"
            else:
                # Fallback: tudo como não encontrado
                for d in docs:
                    doc_result[d] = "nao"

            presentes = [d for d in docs if doc_result[d] == "sim"]
            ausentes  = [d for d in docs if doc_result[d] != "sim"]

            status = "completo"
            if len(ausentes) == len(docs):
                status = "vazio"
            elif len(ausentes) > 0:
                status = "pendente"

            return jsonify({
                "sucesso":    True,
                "nome":       maquina["nome"],
                "url":        maquina.get("url", ""),
                "categoria":  maquina.get("categoria", ""),
                "documentos": doc_result,
                "presentes":  presentes,
                "ausentes":   ausentes,
                "status":     status
            })

        else:
            return jsonify({"erro": f"Tipo de auditoria desconhecido: {tipo}"}), 400

    except Exception as e:
        print("❌ Erro no /tennix-auditoria:", str(e))
        return jsonify({"erro": f"Erro interno: {str(e)}"}), 500


# ─────────────────────────────────────────────────────────────
# 📊 AUDITORIA — Exportar relatório Excel (.xlsx)
# ─────────────────────────────────────────────────────────────
@app.route("/auditoria-exportar-excel", methods=["POST"])
def auditoria_exportar_excel():
    """
    Recebe os resultados da auditoria em JSON e devolve um arquivo .xlsx
    bem formatado com abas de resumo, detalhe por categoria e lista de pendências.
    """
    if not session.get("logado"):
        return jsonify({"erro": "Não autorizado"}), 401

    try:
        from openpyxl import Workbook
        from openpyxl.styles import (
            Font, PatternFill, Alignment, Border, Side, GradientFill
        )
        from openpyxl.utils import get_column_letter
        from openpyxl.chart import BarChart, Reference
        from openpyxl.chart.series import DataPoint
        import io

        data = request.get_json()
        resultados = data.get("resultados", [])
        docs_verificados = data.get("docs_verificados", [])

        if not resultados:
            return jsonify({"erro": "Nenhum resultado para exportar"}), 400

        DOC_LABELS_PT = {
            "folheto":         "Folheto",
            "guia":            "Guia de Peças de Reposição",
            "manual_pecas":    "Manual de Peças",
            "manual_operador": "Manual do Operador",
            "tabela_parede":   "Tabela de Parede",
        }

        # ── Paleta de cores ──────────────────────────────────────
        COR_VERDE_ESCURO   = "1B5C38"   # cabeçalho principal
        COR_VERDE_MEDIO    = "2D8653"   # cabeçalho secundário
        COR_VERDE_CLARO    = "E8F5EE"   # linha alternada
        COR_BRANCO         = "FFFFFF"
        COR_CINZA_HEADER   = "F2F2F2"
        COR_OK             = "D4EDDA"   # célula com ✓
        COR_FALTA          = "FDECEA"   # célula com ✗
        COR_AMARELO        = "FFF3CD"   # pendente
        COR_TEXTO_OK       = "155724"
        COR_TEXTO_FALTA    = "721C24"
        COR_TEXTO_AMARELO  = "856404"

        def fill(hex_color):
            return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

        def border_thin():
            side = Side(style="thin", color="CCCCCC")
            return Border(left=side, right=side, top=side, bottom=side)

        def header_font(size=11, bold=True, color="FFFFFF"):
            return Font(name="Arial", size=size, bold=bold, color=color)

        def cell_font(size=10, bold=False, color="000000"):
            return Font(name="Arial", size=size, bold=bold, color=color)

        wb = Workbook()

        # ════════════════════════════════════════════════════════
        # ABA 1 — RESUMO EXECUTIVO
        # ════════════════════════════════════════════════════════
        ws_resumo = wb.active
        ws_resumo.title = "Resumo"

        # Título
        ws_resumo.merge_cells("A1:H1")
        ws_resumo["A1"] = "RELATÓRIO DE AUDITORIA DE DOCUMENTAÇÃO — TENNANT"
        ws_resumo["A1"].font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
        ws_resumo["A1"].fill = fill(COR_VERDE_ESCURO)
        ws_resumo["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws_resumo.row_dimensions[1].height = 40

        # Data
        ws_resumo.merge_cells("A2:H2")
        ws_resumo["A2"] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws_resumo["A2"].font = Font(name="Arial", size=10, italic=True, color="FFFFFF")
        ws_resumo["A2"].fill = fill(COR_VERDE_MEDIO)
        ws_resumo["A2"].alignment = Alignment(horizontal="center")
        ws_resumo.row_dimensions[2].height = 20

        ws_resumo.append([])

        # Totais gerais
        total    = len(resultados)
        ok       = sum(1 for r in resultados if r.get("status") == "completo")
        pendente = sum(1 for r in resultados if r.get("status") == "pendente")
        vazio    = sum(1 for r in resultados if r.get("status") in ("vazio", "erro"))
        pct_ok   = round(ok / total * 100, 1) if total else 0

        # Cards de sumário (linha 4)
        cards = [
            ("Total de Máquinas", total,    COR_VERDE_ESCURO, "FFFFFF"),
            ("✅ Completas",       ok,       "1B5C38",        "FFFFFF"),
            ("⚠️ Com Pendências",  pendente, "856404",        "FFFFFF"),
            ("❌ Sem Documentos",  vazio,    "721C24",        "FFFFFF"),
            ("Taxa de Conclusão",  f"{pct_ok}%", "2D8653",   "FFFFFF"),
        ]
        # Linha 4: labels
        for i, (label, _, cor, _) in enumerate(cards):
            col = i * 2 + 1
            ws_resumo.merge_cells(
                start_row=4, start_column=col,
                end_row=4, end_column=col + 1
            )
            c = ws_resumo.cell(row=4, column=col, value=label)
            c.font  = Font(name="Arial", size=10, bold=True, color="FFFFFF")
            c.fill  = fill(cor)
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws_resumo.row_dimensions[4].height = 22

        # Linha 5: valores
        for i, (_, valor, cor, _) in enumerate(cards):
            col = i * 2 + 1
            ws_resumo.merge_cells(
                start_row=5, start_column=col,
                end_row=5, end_column=col + 1
            )
            c = ws_resumo.cell(row=5, column=col, value=valor)
            c.font  = Font(name="Arial", size=20, bold=True, color=cor)
            c.fill  = fill(COR_CINZA_HEADER)
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws_resumo.row_dimensions[5].height = 38

        ws_resumo.append([])
        ws_resumo.append([])

        # Resumo por categoria
        categorias = {}
        for r in resultados:
            cat = r.get("categoria") or "Sem categoria"
            if cat not in categorias:
                categorias[cat] = {"total": 0, "completo": 0, "pendente": 0, "vazio": 0}
            categorias[cat]["total"] += 1
            st = r.get("status", "vazio")
            if st == "completo":
                categorias[cat]["completo"] += 1
            elif st == "pendente":
                categorias[cat]["pendente"] += 1
            else:
                categorias[cat]["vazio"] += 1

        row_cat_inicio = ws_resumo.max_row + 1
        headers_cat = ["Categoria", "Total", "Completas", "Pendentes", "Sem Docs", "% OK"]
        for col_i, h in enumerate(headers_cat, 1):
            c = ws_resumo.cell(row=row_cat_inicio, column=col_i, value=h)
            c.font  = header_font()
            c.fill  = fill(COR_VERDE_MEDIO)
            c.alignment = Alignment(horizontal="center")
            c.border = border_thin()
        ws_resumo.row_dimensions[row_cat_inicio].height = 22

        for linha_i, (cat, nums) in enumerate(sorted(categorias.items())):
            row_n = row_cat_inicio + 1 + linha_i
            bg = COR_VERDE_CLARO if linha_i % 2 == 0 else COR_BRANCO
            pct = round(nums["completo"] / nums["total"] * 100, 1) if nums["total"] else 0
            valores = [cat, nums["total"], nums["completo"], nums["pendente"], nums["vazio"], f"{pct}%"]
            for col_i, val in enumerate(valores, 1):
                c = ws_resumo.cell(row=row_n, column=col_i, value=val)
                c.font   = cell_font()
                c.fill   = fill(bg)
                c.border = border_thin()
                c.alignment = Alignment(horizontal="center" if col_i > 1 else "left")

        # Larguras das colunas
        ws_resumo.column_dimensions["A"].width = 35
        for col_ltr in ["B", "C", "D", "E", "F"]:
            ws_resumo.column_dimensions[col_ltr].width = 14
        for col_ltr in ["G", "H"]:
            ws_resumo.column_dimensions[col_ltr].width = 6

        # ════════════════════════════════════════════════════════
        # ABA 2 — DETALHE COMPLETO
        # ════════════════════════════════════════════════════════
        ws_det = wb.create_sheet("Detalhe Completo")

        # Título
        n_cols = 3 + len(docs_verificados) + 2
        ws_det.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
        ws_det["A1"] = "DETALHE DE DOCUMENTAÇÃO POR MÁQUINA"
        ws_det["A1"].font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
        ws_det["A1"].fill = fill(COR_VERDE_ESCURO)
        ws_det["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws_det.row_dimensions[1].height = 32

        # Cabeçalho
        cabecalho = ["Máquina", "Categoria", "URL do Produto"]
        cabecalho += [DOC_LABELS_PT.get(d, d) for d in docs_verificados]
        cabecalho += ["Status", "Documentos Faltando"]

        for col_i, h in enumerate(cabecalho, 1):
            c = ws_det.cell(row=2, column=col_i, value=h)
            c.font  = header_font(size=10)
            c.fill  = fill(COR_VERDE_ESCURO)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = border_thin()
        ws_det.row_dimensions[2].height = 36

        # Dados — ordenados por categoria, depois nome
        resultados_ord = sorted(resultados, key=lambda r: (r.get("categoria",""), r.get("nome","")))

        STATUS_PT = {
            "completo": "✅ Completo",
            "pendente": "⚠️ Pendente",
            "vazio":    "❌ Sem docs",
            "erro":     "⚠️ Erro",
        }
        STATUS_COR = {
            "completo": COR_OK,
            "pendente": COR_AMARELO,
            "vazio":    COR_FALTA,
            "erro":     COR_AMARELO,
        }

        for linha_i, r in enumerate(resultados_ord):
            row_n = 3 + linha_i
            bg_base = COR_VERDE_CLARO if linha_i % 2 == 0 else COR_BRANCO

            faltando = " | ".join(DOC_LABELS_PT.get(d, d) for d in r.get("ausentes", []))
            status   = r.get("status", "vazio")

            linha_vals = [r.get("nome",""), r.get("categoria",""), r.get("url","")]
            for d in docs_verificados:
                val = r.get("documentos", {}).get(d, "nao")
                linha_vals.append("✓" if val == "sim" else "✗")
            linha_vals += [STATUS_PT.get(status, status), faltando]

            for col_i, val in enumerate(linha_vals, 1):
                c = ws_det.cell(row=row_n, column=col_i, value=val)
                c.font   = cell_font()
                c.border = border_thin()
                c.alignment = Alignment(horizontal="left", vertical="center")

                # Colorir células de documento
                if 4 <= col_i <= 3 + len(docs_verificados):
                    if val == "✓":
                        c.fill = fill(COR_OK)
                        c.font = Font(name="Arial", size=11, bold=True, color=COR_TEXTO_OK)
                        c.alignment = Alignment(horizontal="center")
                    else:
                        c.fill = fill(COR_FALTA)
                        c.font = Font(name="Arial", size=11, bold=True, color=COR_TEXTO_FALTA)
                        c.alignment = Alignment(horizontal="center")
                elif col_i == 3 + len(docs_verificados) + 1:
                    # Coluna de status
                    cor_st = STATUS_COR.get(status, COR_CINZA_HEADER)
                    c.fill = fill(cor_st)
                    c.alignment = Alignment(horizontal="center")
                    if status == "completo":
                        c.font = Font(name="Arial", size=10, bold=True, color=COR_TEXTO_OK)
                    elif status in ("vazio", "erro"):
                        c.font = Font(name="Arial", size=10, bold=True, color=COR_TEXTO_FALTA)
                    else:
                        c.font = Font(name="Arial", size=10, bold=True, color=COR_TEXTO_AMARELO)
                else:
                    c.fill = fill(bg_base)

        # Larguras
        ws_det.column_dimensions["A"].width = 20
        ws_det.column_dimensions["B"].width = 28
        ws_det.column_dimensions["C"].width = 40
        for i, d in enumerate(docs_verificados):
            ws_det.column_dimensions[get_column_letter(4 + i)].width = 18
        col_status  = get_column_letter(4 + len(docs_verificados))
        col_faltando = get_column_letter(5 + len(docs_verificados))
        ws_det.column_dimensions[col_status].width = 16
        ws_det.column_dimensions[col_faltando].width = 50

        ws_det.freeze_panes = "A3"

        # ════════════════════════════════════════════════════════
        # ABA 3 — PENDÊNCIAS (só máquinas com algo faltando)
        # ════════════════════════════════════════════════════════
        ws_pend = wb.create_sheet("Pendências")

        ws_pend.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
        ws_pend["A1"] = "LISTA DE PENDÊNCIAS — DOCUMENTOS EM FALTA"
        ws_pend["A1"].font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
        ws_pend["A1"].fill = fill("8B1A1A")
        ws_pend["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws_pend.row_dimensions[1].height = 32

        cabecalho_pend = ["Máquina", "Categoria", "URL", "Documento Faltando", "Prioridade"]
        for col_i, h in enumerate(cabecalho_pend, 1):
            c = ws_pend.cell(row=2, column=col_i, value=h)
            c.font  = Font(name="Arial", size=10, bold=True, color="FFFFFF")
            c.fill  = fill("8B1A1A")
            c.alignment = Alignment(horizontal="center")
            c.border = border_thin()
        ws_pend.row_dimensions[2].height = 24

        pendencias = []
        for r in resultados_ord:
            for doc in r.get("ausentes", []):
                pendencias.append({
                    "nome":       r.get("nome",""),
                    "categoria":  r.get("categoria",""),
                    "url":        r.get("url",""),
                    "doc":        DOC_LABELS_PT.get(doc, doc),
                    "prioridade": "Alta" if doc in ("manual_operador","folheto") else "Média",
                })

        COR_ALTA  = "FDECEA"
        COR_MEDIA = "FFF3CD"
        for linha_i, p in enumerate(pendencias):
            row_n = 3 + linha_i
            bg = COR_ALTA if p["prioridade"] == "Alta" else COR_MEDIA
            vals = [p["nome"], p["categoria"], p["url"], p["doc"], p["prioridade"]]
            for col_i, val in enumerate(vals, 1):
                c = ws_pend.cell(row=row_n, column=col_i, value=val)
                c.fill   = fill(bg)
                c.font   = cell_font()
                c.border = border_thin()
                c.alignment = Alignment(horizontal="center" if col_i in (5,) else "left")

        ws_pend.column_dimensions["A"].width = 20
        ws_pend.column_dimensions["B"].width = 28
        ws_pend.column_dimensions["C"].width = 40
        ws_pend.column_dimensions["D"].width = 28
        ws_pend.column_dimensions["E"].width = 14
        ws_pend.freeze_panes = "A3"

        # ── Salva em memória e retorna ─────────────────────────
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        nome_arquivo = f"auditoria_tennant_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=nome_arquivo
        )

    except Exception as e:
        print("❌ Erro no /auditoria-exportar-excel:", str(e))
        return jsonify({"erro": f"Erro interno: {str(e)}"}), 500


# ─────────────────────────────────────────────────────────────
# 📅 CRONOGRAMA — tipos de conteúdo reconhecidos pelo CSS
# ─────────────────────────────────────────────────────────────
TIPOS_CONTEUDO = ["produto", "dica", "case", "feriado", "conteudo", "outro"]
 
# 📅 CRONOGRAMA — listar posts do cronograma
@app.route("/cronograma", methods=["GET"])
def get_cronograma():
    """Retorna todos os posts do cronograma.
    Suporta filtros: ?mes=6&ano=2026&tipo=produto
    """
    if not session.get("logado"):
        return jsonify({"erro": "Não autorizado"}), 401
 
    mes  = request.args.get("mes",  type=int)
    ano  = request.args.get("ano",  type=int, default=date.today().year)
    tipo = request.args.get("tipo", "").strip().lower()
 
    caminho = os.path.join(os.getcwd(), "cronograma.json")
    if not os.path.exists(caminho):
        return jsonify({"sucesso": True, "posts": [], "total": 0})
 
    with open(caminho, encoding="utf-8") as f:
        posts = json.load(f)
 
    # Filtra por ano
    posts = [p for p in posts if p.get("ano", ano) == ano]
 
    # Filtra por mês se informado
    if mes:
        posts = [p for p in posts if p.get("mes") == mes]
 
    # Filtra por tipo se informado e válido
    if tipo and tipo in TIPOS_CONTEUDO:
        posts = [p for p in posts if p.get("tipo", "outro") == tipo]
 
    return jsonify({"sucesso": True, "posts": posts, "total": len(posts)})
 
 
# 📅 CRONOGRAMA — salvar / substituir posts
@app.route("/cronograma", methods=["POST"])
def salvar_cronograma():
    """Recebe lista de posts e persiste em cronograma.json.
    Body: { "posts": [ { dia, mes, ano, tema, conteudo, tipo, status }, ... ] }
    """
    if not session.get("logado"):
        return jsonify({"erro": "Não autorizado"}), 401
 
    try:
        data = request.get_json()
        posts = data.get("posts", [])
 
        # Valida e normaliza cada post
        normalizados = []
        for p in posts:
            tipo   = p.get("tipo", "outro").lower()
            status = p.get("status", "pendente").lower()
            normalizados.append({
                "dia":      int(p.get("dia", 1)),
                "mes":      int(p.get("mes", 1)),
                "ano":      int(p.get("ano", date.today().year)),
                "tema":     str(p.get("tema", "")).strip(),
                "conteudo": str(p.get("conteudo", "")).strip(),
                "tipo":     tipo   if tipo   in TIPOS_CONTEUDO          else "outro",
                "status":   status if status in ("publicado", "agendado", "pendente") else "pendente",
            })
 
        caminho = os.path.join(os.getcwd(), "cronograma.json")
        with open(caminho, "w", encoding="utf-8") as f:
            json.dump(normalizados, f, ensure_ascii=False, indent=2)
 
        print(f"📅 Cronograma salvo: {len(normalizados)} posts")
        return jsonify({"sucesso": True, "total": len(normalizados)})
 
    except Exception as e:
        print("❌ Erro no /cronograma POST:", e)
        return jsonify({"sucesso": False, "erro": str(e)}), 500
 
 
# 📅 CRONOGRAMA — atualizar status de um post (publicado / agendado / pendente)
@app.route("/cronograma/status", methods=["PATCH"])
def atualizar_status_post():
    """Body: { "dia": 10, "mes": 6, "ano": 2026, "status": "publicado" }"""
    if not session.get("logado"):
        return jsonify({"erro": "Não autorizado"}), 401
 
    try:
        data   = request.get_json()
        dia    = int(data.get("dia"))
        mes    = int(data.get("mes"))
        ano    = int(data.get("ano", date.today().year))
        status = data.get("status", "pendente").lower()
 
        if status not in ("publicado", "agendado", "pendente"):
            return jsonify({"sucesso": False, "erro": "Status inválido"}), 400
 
        caminho = os.path.join(os.getcwd(), "cronograma.json")
        if not os.path.exists(caminho):
            return jsonify({"sucesso": False, "erro": "Cronograma não encontrado"}), 404
 
        with open(caminho, encoding="utf-8") as f:
            posts = json.load(f)
 
        atualizados = 0
        for p in posts:
            if p.get("dia") == dia and p.get("mes") == mes and p.get("ano") == ano:
                p["status"] = status
                atualizados += 1
 
        with open(caminho, "w", encoding="utf-8") as f:
            json.dump(posts, f, ensure_ascii=False, indent=2)
 
        return jsonify({"sucesso": True, "atualizados": atualizados})
 
    except Exception as e:
        print("❌ Erro no /cronograma/status:", e)
        return jsonify({"sucesso": False, "erro": str(e)}), 500
 
 
# ─────────────────────────────────────────────────────────────
# 📊 MÉTRICAS — painel rápido da sidebar
# ─────────────────────────────────────────────────────────────
@app.route("/metricas", methods=["GET"])
def get_metricas():
    """Retorna métricas agregadas para os cards da sidebar."""
    if not session.get("logado"):
        return jsonify({"erro": "Não autorizado"}), 401
 
    hoje = date.today()
    ano  = hoje.year
 
    # Conta arquivos gerados
    total_arquivos = 0
    try:
        total_arquivos = len([
            f for f in os.listdir(PASTA_RESULTADOS)
            if os.path.isfile(os.path.join(PASTA_RESULTADOS, f))
        ])
    except Exception:
        pass
 
    # Lê cronograma para calcular métricas de posts
    posts_publicados = 0
    posts_agendados  = 0
    posts_total_ano  = 0
    proximo_post     = None
 
    caminho = os.path.join(os.getcwd(), "cronograma.json")
    if os.path.exists(caminho):
        try:
            with open(caminho, encoding="utf-8") as f:
                posts = json.load(f)
 
            posts_ano = [p for p in posts if p.get("ano") == ano]
            posts_total_ano  = len(posts_ano)
            posts_publicados = sum(1 for p in posts_ano if p.get("status") == "publicado")
            posts_agendados  = sum(1 for p in posts_ano if p.get("status") == "agendado")
 
            # Próximo post: menor data >= hoje com status agendado ou pendente
            futuros = [
                p for p in posts_ano
                if p.get("status") in ("agendado", "pendente")
                and date(p.get("ano", ano), p.get("mes", 1), p.get("dia", 1)) >= hoje
            ]
            if futuros:
                futuros.sort(key=lambda p: date(p["ano"], p["mes"], p["dia"]))
                prox = futuros[0]
                data_prox = date(prox["ano"], prox["mes"], prox["dia"])
                delta = (data_prox - hoje).days
                proximo_post = {
                    "dia":    prox["dia"],
                    "mes":    prox["mes"],
                    "ano":    prox["ano"],
                    "tema":   prox.get("tema", ""),
                    "tipo":   prox.get("tipo", "outro"),
                    "status": prox.get("status", "pendente"),
                    "dias_restantes": delta,
                    "data_fmt": data_prox.strftime("%d/%m/%Y"),
                }
        except Exception as e:
            print("⚠️ Erro ao ler cronograma para métricas:", e)
 
    # Progresso anual (% de posts publicados em relação ao total)
    pct_anual = round((posts_publicados / posts_total_ano * 100) if posts_total_ano else 0, 1)
 
    return jsonify({
        "sucesso": True,
        "metricas": {
            "arquivos_gerados": total_arquivos,
            "posts_publicados": posts_publicados,
            "posts_agendados":  posts_agendados,
            "posts_total_ano":  posts_total_ano,
            "progresso_anual_pct": pct_anual,
        },
        "proximo_post": proximo_post,
    })
 
 
# ─────────────────────────────────────────────────────────────
# 🏷️  FILTROS — tipos de conteúdo disponíveis
# ─────────────────────────────────────────────────────────────
@app.route("/tipos-conteudo", methods=["GET"])
def get_tipos_conteudo():
    """Retorna os tipos disponíveis para os filtro-chips do CSS."""
    if not session.get("logado"):
        return jsonify({"erro": "Não autorizado"}), 401
 
    labels = {
        "produto":  {"label": "Produto",   "icone": "fa-box"},
        "dica":     {"label": "Dica",      "icone": "fa-lightbulb"},
        "case":     {"label": "Case",      "icone": "fa-star"},
        "feriado":  {"label": "Feriado",   "icone": "fa-calendar"},
        "conteudo": {"label": "Conteúdo",  "icone": "fa-file-alt"},
        "outro":    {"label": "Outro",     "icone": "fa-tag"},
    }
    return jsonify({"sucesso": True, "tipos": labels})
 
 
# ─────────────────────────────────────────────────────────────
# 🔔 TOAST — endpoint para notificações server-side (opcional)
# Permite que o back-end enfileire toasts a serem exibidos
# na próxima requisição do front-end.
# ─────────────────────────────────────────────────────────────
@app.route("/notificacoes", methods=["GET"])
def get_notificacoes():
    """Retorna notificações pendentes para a sessão e as limpa."""
    if not session.get("logado"):
        return jsonify({"erro": "Não autorizado"}), 401
 
    notifs = session.pop("notificacoes", [])
    return jsonify({"sucesso": True, "notificacoes": notifs})
 
 
def _enfileirar_notificacao(tipo: str, titulo: str, mensagem: str = ""):
    """Helper interno: adiciona toast à sessão Flask.
    tipo: 'sucesso' | 'aviso' | 'erro' | 'info'
    """
    notifs = session.get("notificacoes", [])
    notifs.append({"tipo": tipo, "titulo": titulo, "mensagem": mensagem})
    session["notificacoes"] = notifs
 
 
# ─────────────────────────────────────────────────────────────
# 🕵️  VITOR ESPIÃO — Monitoramento de concorrentes (web search)
# ─────────────────────────────────────────────────────────────

ESPIAO_CONCORRENTES = [
    {"id": 1, "name": "Kärcher",   "url": "https://www.karcher.com/br",      "category": "Limpeza Industrial",    "emoji": "🇩🇪"},
    {"id": 2, "name": "Nilfisk",   "url": "https://www.nilfisk.com/pt-br",   "category": "Limpeza Profissional",  "emoji": "🇩🇰"},
    {"id": 3, "name": "Hako",      "url": "https://www.hako.com/br",         "category": "Máquinas Municipais",   "emoji": "🇩🇪"},
    {"id": 4, "name": "Comac",     "url": "https://www.comac.it",            "category": "Lavadoras de Piso",     "emoji": "🇮🇹"},
    {"id": 5, "name": "Fimap",     "url": "https://www.fimap.com/br",        "category": "Limpeza Sustentável",   "emoji": "🇮🇹"},
    {"id": 6, "name": "IPC Group", "url": "https://www.ipcworldwide.com",    "category": "Equip. de Limpeza",     "emoji": "🌍"},
    {"id": 7, "name": "Alabia",    "url": "https://alabia.com.br",          "category": "Robôs de Limpeza",      "emoji": "🇧🇷"},
    {"id": 8, "name": "PUDU",      "url": "https://www.pudurobotics.com/en", "category": "Robôs de Limpeza",     "emoji": "🇨🇳"},
    {"id": 9, "name": "Kunber",    "url": "https://kunber.com.br",          "category": "Equip. de Limpeza",     "emoji": "🇧🇷"},
]

@app.route("/espiao-escanear", methods=["POST"])
def espiao_escanear():
    """
    Usa OpenAI (gpt-4o-mini) com web_search_preview para
    buscar novos produtos e promoções/preços dos concorrentes.
    Body: { "ids": [1, 2, ...] }  — lista de IDs a escanear (opcional; sem body = todos)
    """
    if not session.get("logado"):
        return jsonify({"erro": "Não autorizado"}), 401

    try:
        import requests as req_lib

        data   = request.get_json() or {}
        ids    = data.get("ids", None)   # None = escanear todos

        api_key = os.getenv("OPENAI_API_KEY", "").strip()
        if not api_key:
            return jsonify({"erro": "OPENAI_API_KEY não configurada no servidor."}), 500

        alvos = ESPIAO_CONCORRENTES
        if ids:
            alvos = [c for c in ESPIAO_CONCORRENTES if c["id"] in ids]

        def chamar_openai_responses(prompt_text):
            resp = req_lib.post(
                "https://api.openai.com/v1/responses",
                headers={
                    "Content-Type":  "application/json",
                    "Authorization": f"Bearer {api_key}",
                },
                json={
                    "model": "gpt-4o",
                    "tools": [{"type": "web_search_preview"}],
                    "input": prompt_text,
                    "max_output_tokens": 1500,
                },
                timeout=60,
            )
            if not resp.ok:
                raise Exception(f"OpenAI error {resp.status_code}: {resp.text[:300]}")
            result = resp.json()
            texto = ""
            for item in result.get("output", []):
                if item.get("type") == "message":
                    for c in item.get("content", []):
                        if c.get("type") == "output_text":
                            texto += c.get("text", "")
            return texto

        def extrair_json_seguro(texto):
            import re
            texto = texto.strip()
            try:
                return json.loads(texto)
            except Exception:
                pass
            texto2 = re.sub(r'```json\s*', '', texto)
            texto2 = re.sub(r'```\s*', '', texto2).strip()
            try:
                return json.loads(texto2)
            except Exception:
                pass
            match = re.search(r'\{[\s\S]*\}', texto)
            if match:
                try:
                    return json.loads(match.group(0))
                except Exception:
                    pass
            return None

        resultados = []

        for concorrente in alvos:
            nome = concorrente["name"]
            url  = concorrente["url"]

            prompt = f"""Você é um analista de inteligência competitiva da Tennant Company Brasil.

Pesquise agora no site {url} e na web informações RECENTES (últimas semanas) sobre a empresa {nome} no mercado brasileiro de limpeza profissional.

Busque especificamente:
1. NOVOS PRODUTOS lançados recentemente (máquinas, equipamentos, acessórios)
2. PROMOÇÕES, descontos, campanhas de preço ou condições especiais de pagamento

Para cada item encontrado, classifique como:
- tipo: "produto" (novo lançamento) ou "preco" (promoção/preço/desconto)
- texto: descrição curta e objetiva em português (máx. 100 chars)
- relevancia: "alta", "media" ou "baixa" (impacto para a Tennant)

Responda APENAS com JSON válido, sem texto antes ou depois:
{{"mudancas": [{{"tipo": "produto", "texto": "...", "relevancia": "alta"}}, {{"tipo": "preco", "texto": "...", "relevancia": "media"}}]}}

Se não encontrar nada relevante nas últimas semanas, retorne:
{{"mudancas": []}}"""

            try:
                texto_resp = chamar_openai_responses(prompt)
                parsed     = extrair_json_seguro(texto_resp)
                mudancas   = parsed.get("mudancas", []) if parsed else []
            except Exception as e:
                print(f"❌ Espião erro em {nome}: {e}")
                mudancas = []

            resultados.append({
                "id":        concorrente["id"],
                "name":      nome,
                "url":       url,
                "category":  concorrente["category"],
                "emoji":     concorrente["emoji"],
                "mudancas":  mudancas,
                "alerts":    len(mudancas),
                "status":    "online",
                "scanned_at": datetime.now().strftime("%d/%m/%Y %H:%M"),
            })

        return jsonify({"sucesso": True, "resultados": resultados})

    except Exception as e:
        print("❌ Erro no /espiao-escanear:", str(e))
        return jsonify({"erro": f"Erro interno: {str(e)}"}), 500


# 🚪 Logout
@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))
 
 
# ▶️ Run
if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
 
 
